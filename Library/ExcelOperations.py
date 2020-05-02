import json
import jsonpath
import requests
import openpyxl

class excelUtility:

    def __init__(self, filePath, sheet_name):
        global workbook, sheet
        workbook = openpyxl.load_workbook(filePath)
        sheet = workbook['create']

    def get_row_count(self):
        rows = sheet.max_row
        return rows

    def get_col_count(self):
        cols = sheet.max_column
        return cols

    def get_keyNames(self):
        cols = sheet.max_column
        li=[]
        for i in range(1, cols+1):
            cell = sheet.cell(row=1, column=i)
            li.insert(i-1, cell.value)
        return li


    def update_json_with_data(self, rowNumber, json_request, keyList):
        cols = sheet.max_column
        for i in range(1, cols+1):
            cell = sheet.cell(row= rowNumber, column= i)
            json_request[keyList[i-1]] = cell.value

        return json_request


    def getResourceID(self, rowNumber):
        cell = sheet.cell(row= rowNumber, column=1)
        return cell.value